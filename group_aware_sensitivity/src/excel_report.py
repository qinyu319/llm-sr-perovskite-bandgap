from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import PROJECT_ROOT


SHEET_SOURCES = [
    ("Data_QC", "outputs/summary_tables/task1_data_check.csv"),
    ("Split_Manifest", "outputs/splits/split_manifest.csv"),
    ("Per_Split_Results", "outputs/per_split_results/group_aware_full_workflow_results.csv"),
    ("Term_Frequency", "outputs/summary_tables/group_aware_term_frequency.csv"),
    ("Performance_Summary", "outputs/summary_tables/group_aware_summary.csv"),
    ("Heldout_Group_Error", "outputs/summary_tables/group_aware_heldout_group_summary.csv"),
    ("Selected_Formulas", "outputs/summary_tables/group_aware_selected_formulas.csv"),
    ("Training_Screening", "outputs/per_split_results/training_only_screening.csv"),
]


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(v) else v for v in row])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(ws.iter_cols(), 1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col[:200])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_excel_report(decision_text: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Group-aware Split Sensitivity"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A3"] = "Data scope"
    ws["B3"] = "Original training set only after strict closure QC; original external test excluded from selection."
    ws["A4"] = "Leakage rule"
    ws["B4"] = "All screening, pruning, CV, and formula selection use current outer-train rows only."
    ws["A5"] = "Decision"
    ws["B5"] = decision_text
    ws["A7"] = "Key output"
    ws["B7"] = "See Performance_Summary, Term_Frequency, Heldout_Group_Error, and Selected_Formulas."
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows(min_row=3, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ["A3", "A4", "A5", "A7"]:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill("solid", fgColor="D9EAF7")

    for sheet_name, rel_path in SHEET_SOURCES:
        df = pd.read_csv(PROJECT_ROOT / rel_path)
        target = wb.create_sheet(sheet_name)
        _write_dataframe(target, df)
        if sheet_name == "Per_Split_Results" and target.max_row > 2:
            headers = {c.value: c.column for c in target[1]}
            col = headers.get("test_rmse")
            if col:
                rng = f"{get_column_letter(col)}2:{get_column_letter(col)}{target.max_row}"
                target.conditional_formatting.add(
                    rng,
                    ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
                )

    fig_ws = wb.create_sheet("Figures")
    fig_ws["A1"] = "Publication-ready figures"
    fig_ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    figure_names = [
        "group_aware_rmse_distribution.png",
        "group_aware_term_frequency_heatmap.png",
        "group_aware_formula_similarity.png",
        "group_aware_heldout_group_error.png",
    ]
    anchors = ["A3", "J3", "A31", "J31"]
    for name, anchor in zip(figure_names, anchors):
        path = PROJECT_ROOT / "outputs" / "figures" / name
        if path.exists():
            img = Image(path)
            img.width = 660
            img.height = 420
            fig_ws.add_image(img, anchor)

    out = PROJECT_ROOT / "group_aware_split_sensitivity_results.xlsx"
    wb.save(out)
    return out
